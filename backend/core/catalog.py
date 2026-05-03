"""
catalog.py
----------
Cálculo automático del Catálogo de Conceptos (cantidades de obra)
a partir de los parámetros del proyecto.

Unidades: m, m², m³, pza  (Sistema Internacional)
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import List, Dict


@dataclass
class Concepto:
    clave:      str
    descripcion: str
    unidad:     str
    cantidad:   float
    pu:         float = 0.0   # precio unitario (puede llenarse después)

    @property
    def importe(self) -> float:
        return round(self.cantidad * self.pu, 2)


@dataclass
class CatalogoConceptos:
    conceptos: List[Concepto] = field(default_factory=list)

    def agregar(self, clave: str, desc: str, unidad: str, cantidad: float,
                pu: float = 0.0) -> None:
        self.conceptos.append(Concepto(clave, desc, unidad, round(cantidad, 3), pu))

    def total(self) -> float:
        return sum(c.importe for c in self.conceptos)

    def to_rows(self) -> List[Dict]:
        return [
            {
                "Clave": c.clave,
                "Descripción": c.descripcion,
                "Unidad": c.unidad,
                "Cantidad": c.cantidad,
                "P.U. ($)": c.pu if c.pu else "—",
                "Importe ($)": c.importe if c.pu else "—",
            }
            for c in self.conceptos
        ]

    def to_excel_bytes(self, project_name: str = "Proyecto") -> bytes:
        """Genera un archivo Excel en memoria y devuelve los bytes."""
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catálogo de Conceptos"

        # ---- Encabezado del proyecto ----
        ws.merge_cells("A1:F1")
        ws["A1"] = f"CATÁLOGO DE CONCEPTOS — {project_name.upper()}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        ws["A2"] = "Generado bajo NTC-RCDF  |  lucadstudio.com"
        ws["A2"].font = Font(italic=True, size=9, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        # ---- Cabeceras ----
        headers = ["Clave", "Descripción", "Unidad", "Cantidad", "P.U. ($)", "Importe ($)"]
        header_row = 4
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # ---- Datos ----
        for i, row in enumerate(self.to_rows()):
            r = header_row + 1 + i
            fill = PatternFill("solid", fgColor="EBF3FB" if i % 2 == 0 else "FFFFFF")
            for col, key in enumerate(headers, 1):
                cell = ws.cell(row=r, column=col, value=row[key])
                cell.fill = fill
                cell.border = border
                if col in (4, 5, 6):
                    cell.alignment = Alignment(horizontal="right")

        # ---- Totales ----
        total_row = header_row + 1 + len(self.conceptos) + 1
        ws.merge_cells(f"A{total_row}:E{total_row}")
        ws[f"A{total_row}"] = "TOTAL"
        ws[f"A{total_row}"].font = Font(bold=True)
        ws[f"A{total_row}"].alignment = Alignment(horizontal="right")
        t = self.total()
        ws[f"F{total_row}"] = t if t > 0 else "—"
        ws[f"F{total_row}"].font = Font(bold=True)

        # ---- Ancho de columnas ----
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 48
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
